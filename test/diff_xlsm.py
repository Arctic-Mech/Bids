#!/usr/bin/env python3
"""Diff two xlsm files cell-by-cell: formulas (openpyxl non-data) and values (data_only via original cache).
Usage: diff_xlsm.py a.xlsm b.xlsm"""
import sys
import openpyxl

a_path, b_path = sys.argv[1], sys.argv[2]

def load(path, data_only):
    return openpyxl.load_workbook(path, data_only=data_only, keep_vba=True)

fa, fb = load(a_path, False), load(b_path, False)
fdiffs, vdiffs = [], []

for name in fa.sheetnames:
    if name not in fb.sheetnames:
        fdiffs.append((name, '(sheet missing)', '', ''))
        continue
    wa, wb = fa[name], fb[name]
    maxr = max(wa.max_row, wb.max_row)
    maxc = max(wa.max_column, wb.max_column)
    for row_a, row_b in zip(wa.iter_rows(min_row=1, max_row=maxr, max_col=maxc),
                            wb.iter_rows(min_row=1, max_row=maxr, max_col=maxc)):
        for ca, cb in zip(row_a, row_b):
            va, vb_ = ca.value, cb.value
            if va is None and vb_ is None:
                continue
            fa_is_f = isinstance(va, str) and va.startswith('=')
            fb_is_f = isinstance(vb_, str) and vb_.startswith('=')
            if fa_is_f or fb_is_f:
                if va != vb_:
                    fdiffs.append((name, ca.coordinate, repr(va)[:80], repr(vb_)[:80]))
            else:
                # values: numeric tolerance, string exact
                if isinstance(va, (int, float)) and isinstance(vb_, (int, float)):
                    if abs(float(va) - float(vb_)) > 1e-9:
                        vdiffs.append((name, ca.coordinate, repr(va)[:60], repr(vb_)[:60]))
                elif va != vb_:
                    vdiffs.append((name, ca.coordinate, repr(va)[:60], repr(vb_)[:60]))

print(f"FORMULA differences: {len(fdiffs)}")
for d in fdiffs[:60]:
    print("  F", d)
print(f"VALUE differences: {len(vdiffs)}")
for d in vdiffs[:80]:
    print("  V", d)
